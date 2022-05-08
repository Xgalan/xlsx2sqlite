# -*- coding: utf-8 -*-
"""Utilities for importing or exporting xlsx files.

The :func:`export_worksheet` function creates a xlsx file with a worksheet
containing table data, it uses the `openpyxl` package.

The :func:`import_worksheets` function is useful for importing specified
worksheets from a xlsx file, it uses the `openpyxl` package.
"""
from openpyxl import Workbook, load_workbook

__all__ = ["export_worksheet", "import_worksheet"]


def export_worksheet(filename=None, ws_name=None, rows=None):
    """Export data as a xlsx worksheet file.

    :key filename: Full path of the xlsx file to be saved.
    :key ws_name: Name of the worksheet.
    :key rows: Iterable containing rows data to append to the worksheet.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=ws_name)
    ws.append(rows.headers)
    [ws.append(row) for row in rows]
    wb.save(filename)


def import_worksheet(workbook=None, worksheet=None):
    """Import worksheet from a workbook using openpyxl.

    :key workbook: Name of the xlsx file to open read only.
    :key worksheet: Names of the worksheet to import.

    :returns: A representation of a table with data from the
              imported worksheet.
    """
    # load a xlsx file.
    wb = load_workbook(filename=workbook, read_only=True, keep_vba=False)
    imported_worksheet = wb[worksheet]
    _reset_dimensions(imported_worksheet)
    # import table from imported worksheet
    table = {
        imported_worksheet.title: [
            tuple([cell.value for cell in row]) for row in imported_worksheet.rows
        ]
    }
    wb.close()
    return table


def _reset_dimensions(worksheet):
    """Necessary as to not import a huge amount of empty cells.
    Call appropriate methods from openpyxl to set the correct dimensions of
    the worksheets.
    """
    worksheet.reset_dimensions()
    worksheet.calculate_dimension(force=True)
