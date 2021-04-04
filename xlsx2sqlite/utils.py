# -*- coding: utf-8 -*-
"""Utilities for the command line interface.

The :func:`export_worksheet` function creates a xlsx file with a worksheet
containing table data, it uses the `openpyxl` package.

The :func:`import_worksheets` function is useful for importing specified
worksheets from a xlsx file, it uses the `openpyxl` package.

"""
import configparser

from openpyxl import load_workbook, Workbook


def export_worksheet(filename=None, ws_name=None, rows=None):
    """Export data as a xlsx worksheet file.

    :key filename: Full path of the xlsx file to be saved.
    :key ws_name: Name of the worksheet.
    :key rows: Iterable containing rows data to append to the worksheet.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=ws_name)
    ws.append(rows.headers)
    [ws.append(row) for row in rows]
    wb.save(filename)


def import_worksheets(workbook=None, worksheets=None):
    """Import worksheets from a workbook using openpyxl.

    :key workbook: Name of the xlsx file to open read only.
    :key worksheets: Names of the worksheets to import.

    :returns: A representation of a table with data from the
              imported worksheets.
    """
    # load a xlsx file.
    wb = load_workbook(filename=workbook, read_only=True, keep_vba=False)
    if worksheets is not None:
        imported_worksheets = [wb[ws] for ws in worksheets]
    else:
        imported_worksheets = [wb[ws] for ws in wb.sheetnames]
    # import tables from imported worksheets
    # exclude row that are None ?
    tables = {
        ws.title: [tuple([cell.value for cell in row]) for row in ws.rows] for ws in imported_worksheets
    }
    wb.close()
    return tables


class ConfigModel:
    """Representation for accessing the options of the parsed
    configuration file."""
    COMMA_DELIM = ','
    
    def __init__(self, options=None):
        self.options = options
        self._parser = configparser.ConfigParser()

    def __iter__(self):
        for option in self.options:
            yield option

    def sections(self):
        """List all the sections parsed from the INI file.

        :returns: A list of all the sections declared in the INI file.
        :rtype: list
        """
        return [section for section in self.options]

    def get(self, option):
        """Retrieve a specific option declared in the INI file.

        :param option: Name of the option to retrieve.

        :returns: The value of the option.
        :rtype: str
        """
        try:
            if option in self.options:
                return self.options[option]
            else:
                return [v[option] for k,v in self.options.items() if option in v][0]
        except KeyError as e:
            raise KeyError((str(e) + " not in the options list."))

    def get_imports(self):
        """Retrieve the worksheets names and a subset of columns as declared.

        :returns: A dictionary with a list of worksheets names accessing the
                  `worksheets` key; a list of columns names as a representation
                  for the columns to be retrieved from a worksheet accessing
                  the `subset_cols` key of the dictionary.
                  The lists must be declared in the INI configuration file.
        :rtype: dict
        """
        def get_attrs(names, attribute):
            return dict([
                (name, list(self.get(str(name + attribute).lower()).split(
                    self.COMMA_DELIM))) for name in names
            ])
        names = list(self.get('names').split(self.COMMA_DELIM))
        subset_cols = get_attrs(names, '_columns')
        return {
            'worksheets': names,
            'subset_cols': subset_cols
        }

    def import_config(self, ini):
        """Parse the configuration declared in the INI file.

        Parse the INI file using the `configparser` module, then creates a
        dictionary with all the parsed options.

        :param ini: The path of the INI configuration file to parse.

        :returns: All the options retrieved from the INI file.
        :rtype: dict
        """
        self._parser.read(ini)
        self._inipath = ini
        self.options = {k:dict(self._parser[k])
                        for k in self._parser.sections()}
